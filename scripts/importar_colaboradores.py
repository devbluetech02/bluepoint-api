#!/usr/bin/env python3
"""
Script para importar colaboradores do arquivo colaboradores.xlsx para o banco de dados.
Escala do Excel = Jornada no banco (jornada_id).
"""

import os
import re
import unicodedata
from datetime import datetime

import openpyxl
import psycopg2
import bcrypt

# =====================================================
# CONFIGURAÇÕES (use variáveis de ambiente em produção)
# =====================================================
EXCEL_PATH = os.environ.get("EXCEL_PATH", "colaboradores.xlsx")
SENHA_PADRAO = os.environ.get("SENHA_PADRAO_IMPORT", "Bluepoint@123")

DB_CONFIG = {
    "host": os.environ.get("DB_HOST", "localhost"),
    "port": int(os.environ.get("DB_PORT", "5432")),
    "user": os.environ.get("DB_USERNAME", "bluepoint"),
    "password": os.environ.get("DB_PASSWORD", ""),
    "database": os.environ.get("DB_DATABASE", "bluepoint"),
}

# =====================================================
# FUNÇÕES AUXILIARES
# =====================================================

def normalizar_texto(texto: str) -> str:
    if not texto:
        return ""
    nfkd = unicodedata.normalize('NFKD', texto)
    sem_acento = "".join(c for c in nfkd if not unicodedata.combining(c))
    return sem_acento.upper().strip()


def limpar_cpf(cpf: str) -> str:
    if not cpf:
        return ""
    return re.sub(r'[^\d]', '', cpf)


def valor_ou_none(valor):
    if valor is None:
        return None
    s = str(valor).strip()
    if s in ('--', '', 'None'):
        return None
    return s


def parsear_data(valor) -> str | None:
    if valor is None:
        return None
    s = str(valor).strip()
    if s in ('--', '', 'None'):
        return None
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d')
    try:
        return datetime.strptime(s, '%d/%m/%Y').strftime('%Y-%m-%d')
    except ValueError:
        pass
    try:
        return datetime.strptime(s, '%Y-%m-%d').strftime('%Y-%m-%d')
    except ValueError:
        pass
    print(f"  [WARN] Data não parseada: '{s}'")
    return None


def gerar_email_placeholder(nome: str, cpf_limpo: str) -> str:
    nome_limpo = normalizar_texto(nome).lower().replace(' ', '.')
    nome_limpo = re.sub(r'[^a-z0-9.]', '', nome_limpo)
    sufixo = cpf_limpo[-4:] if len(cpf_limpo) >= 4 else cpf_limpo
    return f"{nome_limpo}.{sufixo}@placeholder.bluepoint.com"


def mapear_status(status_excel: str) -> str:
    if not status_excel:
        return 'ativo'
    s = status_excel.strip().lower()
    if s in ('ativo', 'férias', 'afastado', 'aguardando ativação'):
        return 'ativo'
    return 'inativo'


def mapear_categoria(categoria_excel: str) -> str | None:
    if not categoria_excel:
        return None
    s = str(categoria_excel).strip().lower()
    if 'clt' in s or 'empregado' in s:
        return 'empregado_clt'
    if 'sem categoria' in s or 'gestor' in s:
        return 'usuario_interno'
    return 'empregado_clt'


def hash_senha(senha: str) -> str:
    return bcrypt.hashpw(senha.encode('utf-8'), bcrypt.gensalt(10)).decode('utf-8')


# =====================================================
# MAIN
# =====================================================

def main():
    print("=" * 60)
    print("  IMPORTAÇÃO DE COLABORADORES - BluePoint")
    print("=" * 60)

    # Carregar Excel
    print(f"\nCarregando Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    total_linhas = ws.max_row - 1
    print(f"Total de linhas no Excel: {total_linhas}")

    # Gerar hash da senha
    print("Gerando hash da senha padrão...")
    senha_hash = hash_senha(SENHA_PADRAO)

    # Conectar
    print("\nConectando ao banco de dados...")
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute("SET search_path TO bluepoint, public")

    # =========================================================
    # FASE 1: Coletar dados únicos do Excel e criar referências
    # =========================================================
    print("\n--- FASE 1: Criando departamentos, jornadas e empresas ---\n")

    excel_departamentos = set()
    excel_jornadas = set()
    excel_empresas = set()
    excel_cargos = set()

    for row_num in range(2, ws.max_row + 1):
        dept = valor_ou_none(ws.cell(row=row_num, column=22).value)
        escala = valor_ou_none(ws.cell(row=row_num, column=20).value)
        empresa = valor_ou_none(ws.cell(row=row_num, column=23).value)
        cargo = valor_ou_none(ws.cell(row=row_num, column=21).value)
        if dept:
            excel_departamentos.add(dept.strip())
        if escala:
            excel_jornadas.add(escala.strip())
        if empresa:
            excel_empresas.add(empresa.strip())
        if cargo:
            excel_cargos.add(cargo.strip())

    # --- Departamentos ---
    cur.execute("SELECT id, nome FROM bluepoint.bt_departamentos")
    db_departamentos = {normalizar_texto(row[1]): row[0] for row in cur.fetchall()}
    dept_map = {}  # nome_original -> id

    for dept in sorted(excel_departamentos):
        norm = normalizar_texto(dept)
        if norm in db_departamentos:
            dept_map[dept] = db_departamentos[norm]
        else:
            cur.execute(
                "INSERT INTO bluepoint.bt_departamentos (nome, status) VALUES (%s, 'ativo') RETURNING id",
                (dept,)
            )
            new_id = cur.fetchone()[0]
            db_departamentos[norm] = new_id
            dept_map[dept] = new_id
            print(f"  [NEW] Departamento: '{dept}' (id={new_id})")

    # --- Jornadas ---
    cur.execute("SELECT id, nome FROM bluepoint.bt_jornadas")
    db_jornadas = {normalizar_texto(row[1]): row[0] for row in cur.fetchall()}
    jornada_map = {}  # nome_original -> id

    for jornada in sorted(excel_jornadas):
        norm = normalizar_texto(jornada)
        if norm in db_jornadas:
            jornada_map[jornada] = db_jornadas[norm]
        else:
            cur.execute(
                "INSERT INTO bluepoint.bt_jornadas (nome, status) VALUES (%s, 'ativo') RETURNING id",
                (jornada,)
            )
            new_id = cur.fetchone()[0]
            db_jornadas[norm] = new_id
            jornada_map[jornada] = new_id
            print(f"  [NEW] Jornada: '{jornada}' (id={new_id})")

    # --- Empresas ---
    cur.execute("SELECT id, razao_social FROM bluepoint.bt_empresas ORDER BY id")
    db_empresas_rows = cur.fetchall()
    # Map: razao_social normalizada -> menor ID (primeira ocorrência)
    db_empresas = {}
    for row in db_empresas_rows:
        norm = normalizar_texto(row[1])
        if norm not in db_empresas:
            db_empresas[norm] = row[0]
    empresa_map = {}  # nome_original -> id

    for empresa in sorted(excel_empresas):
        norm = normalizar_texto(empresa)
        if norm in db_empresas:
            empresa_map[empresa] = db_empresas[norm]
        else:
            # Gerar CNPJ placeholder único
            placeholder_cnpj = f"00000000{hash(empresa) % 999999:06d}"
            try:
                cur.execute(
                    "INSERT INTO bluepoint.bt_empresas (razao_social, nome_fantasia, cnpj) VALUES (%s, %s, %s) RETURNING id",
                    (empresa, empresa, placeholder_cnpj)
                )
                new_id = cur.fetchone()[0]
                db_empresas[norm] = new_id
                empresa_map[empresa] = new_id
                print(f"  [NEW] Empresa: '{empresa}' (id={new_id}) - CNPJ placeholder, atualizar!")
            except Exception as e:
                conn.rollback()
                cur.execute("SET search_path TO bluepoint, public")
                print(f"  [ERROR] Empresa '{empresa}': {e}")
                # Tentar com outro CNPJ
                placeholder_cnpj2 = f"99{abs(hash(empresa + '2')) % 99999999:08d}{abs(hash(empresa)) % 99:02d}"
                try:
                    cur.execute(
                        "INSERT INTO bluepoint.bt_empresas (razao_social, nome_fantasia, cnpj) VALUES (%s, %s, %s) RETURNING id",
                        (empresa, empresa, placeholder_cnpj2)
                    )
                    new_id = cur.fetchone()[0]
                    db_empresas[norm] = new_id
                    empresa_map[empresa] = new_id
                    print(f"  [RETRY OK] Empresa: '{empresa}' (id={new_id})")
                except Exception as e2:
                    conn.rollback()
                    cur.execute("SET search_path TO bluepoint, public")
                    print(f"  [FATAL] Empresa '{empresa}' não pôde ser criada: {e2}")

    # --- Cargos ---
    cur.execute("SELECT id, nome FROM bluepoint.bt_cargos")
    db_cargos = {normalizar_texto(row[1]): row[0] for row in cur.fetchall()}
    cargo_map = {}  # nome_original -> id

    for cargo in sorted(excel_cargos):
        norm = normalizar_texto(cargo)
        if norm in db_cargos:
            cargo_map[cargo] = db_cargos[norm]
        else:
            cur.execute(
                "INSERT INTO bluepoint.bt_cargos (nome) VALUES (%s) RETURNING id",
                (cargo,)
            )
            new_id = cur.fetchone()[0]
            db_cargos[norm] = new_id
            cargo_map[cargo] = new_id
            print(f"  [NEW] Cargo: '{cargo}' (id={new_id})")

    # Commit das referências
    conn.commit()
    print(f"\nReferências criadas e commitadas!")
    print(f"  Departamentos mapeados: {len(dept_map)}")
    print(f"  Jornadas mapeadas: {len(jornada_map)}")
    print(f"  Empresas mapeadas: {len(empresa_map)}")
    print(f"  Cargos mapeados: {len(cargo_map)}")

    # =========================================================
    # FASE 2: Inserir colaboradores
    # =========================================================
    print(f"\n--- FASE 2: Inserindo colaboradores ---\n")

    # Buscar CPFs existentes
    cur.execute("SELECT cpf FROM bluepoint.bt_colaboradores")
    cpfs_existentes = set(row[0] for row in cur.fetchall())
    print(f"CPFs já no banco: {len(cpfs_existentes)}")

    # Buscar emails existentes
    cur.execute("SELECT lower(email) FROM bluepoint.bt_colaboradores")
    emails_existentes = set(row[0] for row in cur.fetchall())

    inseridos = 0
    pulados_cpf = 0
    erros = 0

    for row_num in range(2, ws.max_row + 1):
        nome = valor_ou_none(ws.cell(row=row_num, column=1).value)
        cpf_raw = valor_ou_none(ws.cell(row=row_num, column=2).value)
        pis = valor_ou_none(ws.cell(row=row_num, column=3).value)
        nascimento_raw = ws.cell(row=row_num, column=4).value
        email_raw = valor_ou_none(ws.cell(row=row_num, column=5).value)
        celular = valor_ou_none(ws.cell(row=row_num, column=6).value)
        telefone = valor_ou_none(ws.cell(row=row_num, column=7).value)
        observacoes = valor_ou_none(ws.cell(row=row_num, column=8).value)
        logradouro = valor_ou_none(ws.cell(row=row_num, column=9).value)
        numero = valor_ou_none(ws.cell(row=row_num, column=10).value)
        complemento = valor_ou_none(ws.cell(row=row_num, column=11).value)
        bairro = valor_ou_none(ws.cell(row=row_num, column=12).value)
        cidade = valor_ou_none(ws.cell(row=row_num, column=13).value)
        estado = valor_ou_none(ws.cell(row=row_num, column=14).value)
        cep = valor_ou_none(ws.cell(row=row_num, column=15).value)
        status_excel = valor_ou_none(ws.cell(row=row_num, column=17).value)
        admissao_raw = ws.cell(row=row_num, column=18).value
        desligamento_raw = ws.cell(row=row_num, column=19).value
        escala = valor_ou_none(ws.cell(row=row_num, column=20).value)
        cargo = valor_ou_none(ws.cell(row=row_num, column=21).value)
        departamento_nome = valor_ou_none(ws.cell(row=row_num, column=22).value)
        empresa_nome = valor_ou_none(ws.cell(row=row_num, column=23).value)
        categoria_excel = valor_ou_none(ws.cell(row=row_num, column=24).value)

        if not nome or not cpf_raw:
            erros += 1
            continue

        cpf_limpo = limpar_cpf(cpf_raw)

        # Pular CPFs duplicados
        if cpf_limpo in cpfs_existentes:
            pulados_cpf += 1
            continue

        # Email
        email = email_raw.lower().strip() if email_raw else None
        if not email or '@' not in email:
            email = gerar_email_placeholder(nome, cpf_limpo)

        if email.lower() in emails_existentes:
            base, dominio = email.rsplit('@', 1)
            cnt = 1
            while f"{base}.{cnt}@{dominio}".lower() in emails_existentes:
                cnt += 1
            email = f"{base}.{cnt}@{dominio}"

        # Datas
        data_nascimento = parsear_data(nascimento_raw)
        data_admissao = parsear_data(admissao_raw)
        data_desligamento = parsear_data(desligamento_raw)
        if not data_admissao:
            data_admissao = '2024-01-01'

        status = mapear_status(status_excel)
        categoria = mapear_categoria(categoria_excel)
        tel = celular or telefone

        # IDs de referência
        departamento_id = dept_map.get(departamento_nome.strip()) if departamento_nome else None
        jornada_id = jornada_map.get(escala.strip()) if escala else None
        empresa_id = empresa_map.get(empresa_nome.strip()) if empresa_nome else None
        cargo_id = cargo_map.get(cargo.strip()) if cargo else None

        try:
            # Usar SAVEPOINT para não perder o que já inseriu
            cur.execute("SAVEPOINT sp_colab")
            cur.execute("""
                INSERT INTO bluepoint.bt_colaboradores (
                    nome, email, senha_hash, cpf, telefone, pis, categoria, observacao,
                    cargo_id, empresa_id, departamento_id, jornada_id,
                    data_admissao, data_nascimento, data_desligamento,
                    status, tipo,
                    endereco_cep, endereco_logradouro, endereco_numero,
                    endereco_complemento, endereco_bairro, endereco_cidade, endereco_estado,
                    permite_ponto_mobile, permite_ponto_qualquer_empresa
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, 'colaborador',
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    false, false
                )
            """, (
                nome.strip(),
                email,
                senha_hash,
                cpf_limpo,
                tel,
                pis,
                categoria,
                observacoes,
                cargo_id,
                empresa_id,
                departamento_id,
                jornada_id,
                data_admissao,
                data_nascimento,
                data_desligamento,
                status,
                cep,
                logradouro,
                str(numero) if numero else None,
                complemento,
                bairro,
                cidade,
                estado,
            ))
            cur.execute("RELEASE SAVEPOINT sp_colab")

            cpfs_existentes.add(cpf_limpo)
            emails_existentes.add(email.lower())
            inseridos += 1

            if inseridos % 100 == 0:
                print(f"  ... {inseridos} colaboradores inseridos ...")

        except Exception as e:
            cur.execute("ROLLBACK TO SAVEPOINT sp_colab")
            print(f"  [ERROR] Linha {row_num} ({nome}): {e}")
            erros += 1

    # Commit final
    conn.commit()

    print(f"\n{'=' * 60}")
    print(f"  RESULTADO DA IMPORTAÇÃO")
    print(f"{'=' * 60}")
    print(f"  Total no Excel:          {total_linhas}")
    print(f"  Inseridos com sucesso:   {inseridos}")
    print(f"  Pulados (CPF duplicado): {pulados_cpf}")
    print(f"  Erros:                   {erros}")
    print(f"{'=' * 60}")
    print(f"  Senha padrão: {SENHA_PADRAO}")
    print(f"{'=' * 60}")

    cur.execute("SELECT count(*) FROM bluepoint.bt_colaboradores")
    total_banco = cur.fetchone()[0]
    print(f"\n  Total de colaboradores no banco agora: {total_banco}")

    cur.close()
    conn.close()
    print("\nConcluído!")


if __name__ == "__main__":
    main()
